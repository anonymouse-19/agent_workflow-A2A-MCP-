"""MCP Tool: Excel Reader — Reads and parses Excel spreadsheet data."""

import os


def read_excel(file_path: str) -> dict:
    """
    Read data from an Excel (.xlsx) file.
    Uses openpyxl (free, open-source). Falls back gracefully if not installed.
    """
    if not os.path.exists(file_path):
        return {"error": f"File not found: {file_path}"}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheets = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row is None or ws.max_row == 0:
                sheets[sheet_name] = {"headers": [], "rows": [], "row_count": 0}
                continue

            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value) if cell.value is not None else "")

            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append([v if v is not None else "" for v in row])

            sheets[sheet_name] = {
                "headers": headers,
                "rows": rows,
                "row_count": len(rows),
            }

        return {
            "sheets": sheets,
            "sheet_names": wb.sheetnames,
            "file": os.path.basename(file_path),
            "type": "excel",
        }
    except ImportError:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}
    except Exception as e:
        return {"error": f"Failed to read Excel: {str(e)}"}
