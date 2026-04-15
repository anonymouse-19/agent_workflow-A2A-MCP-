"""
MCP Server — Exposes all tools via the Model Context Protocol.

Tools: read_pdf, read_excel, read_text, extract_questions,
       summarize_text, extract_keywords, analyze_tabular_data, llm_enhance

Run: python mcp_server/server.py
"""

import os
import sys
import re
import json
from collections import Counter
from typing import Any

from mcp.server.fastmcp import FastMCP

# ── MCP Server Instance ─────────────────────────────────────────────────────

mcp = FastMCP(
    "HealthcareToolServer",
    stateless_http=True,
    json_response=True,
    host="0.0.0.0",
    port=int(os.environ.get("MCP_PORT", "8004")),
)

# ── Stopwords ────────────────────────────────────────────────────────────────

STOP_WORDS = frozenset({
    "the", "and", "for", "are", "but", "not", "you", "all", "can", "had",
    "her", "was", "one", "our", "out", "has", "have", "been", "that", "this",
    "with", "they", "from", "will", "each", "make", "like", "just", "over",
    "such", "than", "them", "some", "into", "its", "also", "more", "other",
    "would", "which", "their", "what", "about", "there", "when", "your",
    "could", "should", "these", "those", "where", "being", "does", "only",
    "most", "very", "much", "then", "here", "after", "before", "while",
})


# ── Tool: read_pdf ───────────────────────────────────────────────────────────

@mcp.tool()
def read_pdf(file_path: str) -> dict:
    """Read and extract text from a PDF file.

    Args:
        file_path: Absolute or relative path to the PDF file.

    Returns:
        Dictionary with content, pages, chars, file name, and type.
    """
    if not os.path.exists(file_path):
        return {"error": f"File not found: {file_path}"}
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(file_path)
        pages_text = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                pages_text.append(text)
        content = "\n\n".join(pages_text)
        return {
            "content": content.strip(),
            "pages": len(reader.pages),
            "chars": len(content),
            "file": os.path.basename(file_path),
            "type": "pdf",
        }
    except ImportError:
        return {"error": "PyPDF2 not installed. Run: pip install PyPDF2"}
    except Exception as e:
        return {"error": f"Failed to read PDF: {str(e)}"}


# ── Tool: read_excel ─────────────────────────────────────────────────────────

@mcp.tool()
def read_excel(file_path: str) -> dict:
    """Read data from an Excel (.xlsx) file.

    Args:
        file_path: Absolute or relative path to the Excel file.

    Returns:
        Dictionary with sheets data, sheet names, file name, and type.
    """
    if not os.path.exists(file_path):
        return {"error": f"File not found: {file_path}"}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheets = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row is None or ws.max_row == 0:
                sheets[sheet_name] = {"headers": [], "rows": [], "row_count": 0}
                continue
            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value) if cell.value is not None else "")
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append([v if v is not None else "" for v in row])
            sheets[sheet_name] = {
                "headers": headers,
                "rows": rows,
                "row_count": len(rows),
            }
        return {
            "sheets": sheets,
            "sheet_names": wb.sheetnames,
            "file": os.path.basename(file_path),
            "type": "excel",
        }
    except ImportError:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}
    except Exception as e:
        return {"error": f"Failed to read Excel: {str(e)}"}


# ── Tool: read_text ──────────────────────────────────────────────────────────

@mcp.tool()
def read_text(file_path: str) -> dict:
    """Read content from a plain text or CSV file.

    Args:
        file_path: Absolute or relative path to the text file.

    Returns:
        Dictionary with content, char count, line count, file name, and type.
    """
    if not os.path.exists(file_path):
        return {"error": f"File not found: {file_path}"}
    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read()
        return {
            "content": content,
            "chars": len(content),
            "lines": content.count("\n") + 1,
            "file": os.path.basename(file_path),
            "type": "text",
        }
    except Exception as e:
        return {"error": f"Failed to read file: {str(e)}"}


# ── Tool: extract_questions ──────────────────────────────────────────────────

@mcp.tool()
def extract_questions(text: str) -> dict:
    """Extract all questions from text content.

    Args:
        text: The text to scan for questions.

    Returns:
        Dictionary with list of questions and count.
    """
    sentences = re.split(r'(?<=[.!?])\s+', text)
    questions = []
    seen: set = set()

    for sent in sentences:
        sent = sent.strip()
        if not sent:
            continue
        if sent.endswith("?") and sent not in seen:
            questions.append(sent)
            seen.add(sent)

    pattern = r'(?:^|\n)\s*(?:\d+[\.\)]\s*)?(?:What|How|Why|When|Where|Who|Which|Can|Do|Does|Is|Are|Will|Should|Could|Would)[^\n.]*\?'
    for match in re.findall(pattern, text, re.IGNORECASE | re.MULTILINE):
        q = match.strip()
        if q not in seen:
            questions.append(q)
            seen.add(q)

    return {"questions": questions, "count": len(questions)}


# ── Tool: summarize_text ─────────────────────────────────────────────────────

@mcp.tool()
def summarize_text(text: str, num_sentences: int = 5) -> dict:
    """Create an extractive summary of text using TF-based sentence scoring.

    Args:
        text: The text to summarize.
        num_sentences: Number of sentences in the summary (default 5).

    Returns:
        Dictionary with summary text, method used, and sentence counts.
    """
    sentences = re.split(r'(?<=[.!?])\s+', text)
    sentences = [s.strip() for s in sentences if len(s.strip()) > 20]

    if not sentences:
        return {"summary": text[:500] if text else "(empty)", "method": "truncation"}
    if len(sentences) <= num_sentences:
        return {"summary": " ".join(sentences), "method": "full_text"}

    words = re.findall(r'\b[a-z]{3,}\b', text.lower())
    word_freq = Counter(w for w in words if w not in STOP_WORDS)

    if not word_freq:
        return {"summary": " ".join(sentences[:num_sentences]), "method": "first_n"}

    max_freq = max(word_freq.values())
    normalized = {w: freq / max_freq for w, freq in word_freq.items()}

    scored = []
    for i, sent in enumerate(sentences):
        words_in = re.findall(r'\b[a-z]{3,}\b', sent.lower())
        if words_in:
            score = sum(normalized.get(w, 0) for w in words_in) / len(words_in)
            scored.append((score, i, sent))

    scored.sort(reverse=True)
    top = scored[:num_sentences]
    top.sort(key=lambda x: x[1])

    summary = " ".join(sent for _, _, sent in top)
    return {
        "summary": summary,
        "method": "extractive_tf",
        "sentences_selected": len(top),
        "total_sentences": len(sentences),
    }


# ── Tool: extract_keywords ──────────────────────────────────────────────────

@mcp.tool()
def extract_keywords(text: str, top_n: int = 10) -> dict:
    """Extract the most important keywords from text using term frequency.

    Args:
        text: The text to analyze.
        top_n: Number of top keywords to return (default 10).

    Returns:
        Dictionary with keywords list, frequencies, and unique word count.
    """
    words = re.findall(r'\b[a-z]{3,}\b', text.lower())
    word_freq = Counter(w for w in words if w not in STOP_WORDS)
    top = word_freq.most_common(top_n)
    return {
        "keywords": [w for w, _ in top],
        "frequencies": dict(top),
        "total_unique_words": len(word_freq),
    }


# ── Tool: analyze_tabular_data ───────────────────────────────────────────────

@mcp.tool()
def analyze_tabular_data(headers: list[str], rows: list[list[Any]]) -> dict:
    """Analyze tabular/spreadsheet data and generate statistical insights.

    Args:
        headers: List of column header strings.
        rows: List of data rows (each row is a list of values).

    Returns:
        Dictionary with row_count, column_count, column_stats, and insights.
    """
    if not rows or not headers:
        return {"error": "No data to analyze"}

    insights = []
    column_stats = {}

    for i, header in enumerate(headers):
        if not header:
            continue
        values = [row[i] for row in rows if i < len(row) and row[i] not in (None, "")]

        numeric_vals = []
        for v in values:
            try:
                numeric_vals.append(float(v))
            except (ValueError, TypeError):
                pass

        if numeric_vals and len(numeric_vals) >= 2:
            sorted_v = sorted(numeric_vals)
            mid = len(sorted_v) // 2
            median = sorted_v[mid] if len(sorted_v) % 2 else (sorted_v[mid - 1] + sorted_v[mid]) / 2
            mean = sum(numeric_vals) / len(numeric_vals)
            variance = sum((x - mean) ** 2 for x in numeric_vals) / len(numeric_vals)
            std_dev = variance ** 0.5

            # Flag anomalies beyond 2 standard deviations
            anomalies = [v for v in numeric_vals if abs(v - mean) > 2 * std_dev]

            stats = {
                "type": "numeric",
                "count": len(numeric_vals),
                "min": min(numeric_vals),
                "max": max(numeric_vals),
                "mean": round(mean, 2),
                "median": round(median, 2),
                "std_dev": round(std_dev, 2),
                "sum": round(sum(numeric_vals), 2),
                "anomalies_count": len(anomalies),
            }
            column_stats[header] = stats
            insights.append(
                f"'{header}': numeric, range [{stats['min']}-{stats['max']}], "
                f"mean={stats['mean']}, median={stats['median']}, std={stats['std_dev']}"
            )
            if anomalies:
                insights.append(
                    f"'{header}': {len(anomalies)} anomalies detected beyond 2 std deviations"
                )

            # Trend detection (simple: compare first half mean vs second half mean)
            if len(numeric_vals) >= 4:
                half = len(numeric_vals) // 2
                first_half_mean = sum(numeric_vals[:half]) / half
                second_half_mean = sum(numeric_vals[half:]) / (len(numeric_vals) - half)
                if second_half_mean > first_half_mean * 1.05:
                    insights.append(f"'{header}': increasing trend detected")
                elif second_half_mean < first_half_mean * 0.95:
                    insights.append(f"'{header}': decreasing trend detected")

        elif values:
            value_counts = Counter(str(v) for v in values)
            most_common = value_counts.most_common(3)
            stats = {
                "type": "categorical",
                "count": len(values),
                "unique": len(value_counts),
                "most_common": most_common,
            }
            column_stats[header] = stats
            top_label = most_common[0][0] if most_common else "N/A"
            insights.append(
                f"'{header}': categorical, {len(value_counts)} unique values, "
                f"most common: '{top_label}'"
            )

    return {
        "row_count": len(rows),
        "column_count": len(headers),
        "column_stats": column_stats,
        "insights": insights,
    }


# ── Tool: llm_enhance ───────────────────────────────────────────────────────

@mcp.tool()
def llm_enhance(text: str, context: str = "", task_type: str = "enhance") -> dict:
    """Enhance text analysis using Groq LLM for clinical interpretation.

    Args:
        text: Text content to enhance or interpret.
        context: Additional context for the LLM.
        task_type: Type of enhancement - 'enhance', 'insights', or 'interpret'.

    Returns:
        Dictionary with enhanced content, model used, and token usage.
    """
    # Import here to avoid circular deps and allow server to start without GROQ_API_KEY
    # for non-LLM tools
    try:
        sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        from shared.llm_adapter import GroqAdapter
        adapter = GroqAdapter()
    except RuntimeError as e:
        return {"error": str(e), "fallback": True}

    system_msg = (
        "You are a clinical data analyst assistant. "
        "Provide concise, evidence-based interpretations. "
        "Do not make diagnostic claims."
    )

    if task_type == "insights":
        user_msg = (
            f"Given the following clinical data statistics, provide 3-5 clinical insights "
            f"and actionable recommendations:\n\nContext: {context}\n\nData:\n{text}"
        )
    elif task_type == "interpret":
        user_msg = (
            f"Provide a natural language interpretation of this clinical analysis:\n\n"
            f"Context: {context}\n\nAnalysis:\n{text}"
        )
    else:
        user_msg = (
            f"Improve this clinical summary for clarity and add clinical implications.\n\n"
            f"Context: {context}\n\nSummary: {text}\n\nEnhanced summary:"
        )

    result = adapter.complete(
        messages=[
            {"role": "system", "content": system_msg},
            {"role": "user", "content": user_msg},
        ],
        temperature=0.3,
    )

    return {
        "enhanced_text": result["content"],
        "model": result["model"],
        "input_tokens": result["input_tokens"],
        "output_tokens": result["output_tokens"],
        "latency_ms": result["latency_ms"],
    }


# ── Main ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("Starting MCP Healthcare Tool Server on port", os.environ.get("MCP_PORT", "8004"))
    mcp.run(transport="streamable-http")
